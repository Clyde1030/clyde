import os
from datetime import datetime
from dateutil import parser
import xlwings as xw
import openpyxl as xl
from openpyxl.utils.cell import get_column_letter

def find_last_col_row(path, sheet_name):
    '''find the last row for a worksheet'''
    with xw.App(visible=False) as app:
        wb = xw.Book(path)
        sht = wb.sheets[sheet_name]
        last_col = sht.range('A1').end('right').column
        last_col_letter = get_column_letter(last_col)
        last_row = sht.range('A1').end('down').row
    print(r"{} tab of {}'s last column is {}, and its last row number is {}".format(sheet_name,os.path.basename(path),last_col_letter,last_row))
    return last_col, last_row

def find_last_row(path, sheet_name):
    '''find the last row for a worksheet'''
    wb = xl.load_workbook(path)
    ws = wb[sheet_name]
    last_row = len(ws['A1'])
    wb.close()
    print(r"{} tab of {}'s last row number is {}".format(sheet_name,os.path.basename(path),last_row))
    return last_row

def pasteRange(startCol,startRow,endCol,endRow,sheetReceiving,copiedData):
    """
    Paste a list to a specific range of an excel worksheet. 
    Note that copiedData has to be a list! 
    Use df.values.tolist() method to convert dataframe to a list if needed.
    """
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

def formatDate(date):
    asDate = parser.parse(date)
    asString = asDate.strftime('%Y-%m-%d %X')
    return asString

def formatMoney(number):
    asFloat = round(float(number),4)
    asString = '{:.4f}'.format(asFloat)
    return asString

def formatQS(number):
    asFloat = round(float(number),8)
    asString = '{:.8f}'.format(asFloat)
    return asString

def formatRating(number):
    asFloat = round(float(number),7)
    asString = '{:.7f}'.format(asFloat)
    return asString

# Debugging Area:
# def main():
#     col, row = find_last_col_row(r'C:\Users\ylee\Desktop\dev\AUL\MONTHLYREPORT\etl\Yield Solutions Claims Detail Append.xlsx','data')
#     wb = openwb2(r'C:\Users\ylee\Desktop\dev\AUL\MONTHLYREPORT\etl\Yield Solutions Claims Detail Append.xlsx')
# if __name__ == '__main__':
#     main()



