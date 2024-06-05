import os
import shutil
import logging
import re
import datetime
import time
import math

import openpyxl as xl
import pandas as pd
from sqlalchemy import create_engine, text


# LOGGER = logging.getLogger(__name__)
LOGGER = logging.getLogger()
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
handler = logging.StreamHandler()
handler.setFormatter(formatter)
LOGGER.addHandler(handler)
LOGGER.setLevel(logging.DEBUG)

class letters:

    def __init__(self, year, quarter):
        self.year = year
        self.quarter = quarter
        self.Qquarter = f'Q{quarter}'
        self.maindir = f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers'
        self.localdir = os.getcwd()
        self.ad = year*100+quarter*3
        if quarter == 1:
            self.prioryear = year -1
        else:
            self.priorquarter = 4
            self.prioryear = year
            self.priorquarter = quarter-1


    def do_ace(self):

        priorQfile = os.path.join(self.maindir,str(self.prioryear),f'{self.prioryear} Q{self.priorquarter} Letters','ACE',f'{self.prioryear}Q{self.priorquarter} ACE Cessions and Receivables.xlsx')
        Qfile = os.path.join(self.maindir,str(self.year),f'{self.year} Q{self.quarter} Letters','ACE',f'{self.year}Q{self.quarter} ACE Cessions and Receivables.xlsx')
        # Qfile = os.path.join(self.localdir,f'{self.year}Q{self.quarter} ACE Cessions and Receivables.xlsx')
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")

        if not os.path.exists(Qfile):
            shutil.copy(priorQfile,Qfile)
            LOGGER.info(f'{os.path.basename(Qfile)} is created at {os.path.dirname(Qfile)}')

        qry = '''
            select policynumber, issuedate, reinsurancetreaty, assumedface, retainedface, cedednar 			
            from inforce			
            where reinsurercompany = 'at' and activitydate = '''+str(self.ad)+''' and isjoint in ('n', 'p')			
            order by cedednar desc			
            '''
        
        data, rowcount, headers = self._getdata('pdxvmdb11','TAI',qry) # return a list

        ace = xl.load_workbook(Qfile)
        ws = ace.active

        # clear values in A22:F5000
        for row in ws['A22:F5000']:
            for cell in row:
                cell.value = None

        # Update Cells formula and values to point to new quarter
        ws.cell(20,1).value = f'Q{self.quarter} {self.year}'
        ws.cell(8,7).value = f"='J:\Acctng\QuarterClose\{self.year}\Q{self.quarter}\STAT\[{self.year}Q{self.quarter} IBNR.xlsx]IBNR'!$F$172"
        ws.cell(8,8).value = f"='J:\Acctng\QuarterClose\{self.year}\Q{self.quarter}\Data\[{self.year}Q{self.quarter} TAI Queries.xlsm]Exposure net of HLR Modco'!F90"
        ws.cell(20,4).value = f"=SUM(D22:D{rowcount+21})"
        ws.cell(20,5).value = f"=SUM(E22:E{rowcount+21})"
        ws.cell(20,6).value = f"=SUM(F22:F{rowcount+21})"
        for row in range(22, rowcount+22,1):
            ws.cell(row,9).value = f"=F{row}*$I$8"
            ws.cell(row,12).value = f"=(F{row}/0.15+E{row})*$I$8"
            ws.cell(row,13).value = f"=I{row}/0.15"
            ws.cell(row,14).value = f"=E{row}*$I$8"
            ws.cell(row,16).value = f'=IF(M{row}=0, "N/A", I{row}/M{row})'

        # paste query data onto worksheet starting from cell A22
        self._pasteRange(1,22,6,rowcount+21,ws,data.values.tolist())
        
        ace.save(Qfile)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")


    def do_hannover(self):

        # Qfile = os.path.join(self.localdir,f'{self.year}Q{self.quarter} Hannover Life Re Reserves.xlsx')
        Qfile = os.path.join(f'J:\MLife\Reinsurance\Reinsurers\Hannover\TAI Reserves\{self.year}',f'{self.year}Q{self.quarter} Hannover Life Re Reserves.xlsx')
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")

        qry = open(os.path.join(self.localdir,'QuarterlyReservePull - Hannover.sql')).read()
        data, rowcount, headers = self._getdata('pdxvmdb11','TAI',qry, ad=self.ad)
        data.to_excel(Qfile,header = True, index = False)
        path2 = f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.year}\{self.year} Q{self.quarter} Letters\Hannover\HLR'
        shutil.copy(Qfile,path2)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")
        LOGGER.info(f"{os.path.basename(Qfile)} is copied to {path2}.")        


    def do_jh(self):

        Qfile = os.path.join(f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.year}\{self.year} Q{self.quarter} Letters\John Hancock',f'JH {self.year}Q{self.quarter} Majestic Billings by Reinsurer and Plancode.xlsx')
        # Qfile = os.path.join(self.localdir,f'JH {self.year}Q{self.quarter} Majestic Billings by Reinsurer and Plancode.xlsx')
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")        
        
        JHQuery2 = '''									
        select reinsurercompany, cedingcompany, billingTitle, plancode,	sum(totalPremium) as SumPremium										
        from dbo.VBillingSummaryDetail										
        where activityDate in ('''+ str(self.ad) +''','''+str(self.ad-1)+''','''+str(self.ad-2)+''') 
        and cedingcompany = '1jh' and reinsurercompany in ('RG') and (billingtitle like ('%sum of d%') or billingtitle like ('%sum of e%') or billingtitle like ('%sum of f%') or billingtitle like ('%sum of g%'))										
        group by reinsurercompany, cedingcompany, billingTitle, plancode										
        order by plancode										
        '''
        
        JHQuery1 = open(os.path.join(self.localdir,'JH Maj Billing.sql')).read()
        data1, rowcount1, headers1 = self._getdata('pdxvmdb11','TAI',JHQuery1,ad=self.ad)
        data2, rowcount2, headers2  = self._getdata('pdxvmdb11','TAI',JHQuery2)

        wb = xl.Workbook()
        ws = wb.active

        self._pasteRange(1,2,3,1+rowcount1,ws,data1.values.tolist())
        self._pasteRange(6,2,10,1+rowcount2,ws,data2.values.tolist())

        # Fill the column headings
        ws.cell(1,1).value = 'reinsurercompany'
        ws.cell(1,2).value = 'plancode'
        ws.cell(1,3).value = 'SumPremium'
        ws.cell(1,6).value = 'reinsurercompany'
        ws.cell(1,7).value = 'cedingcompany'
        ws.cell(1,8).value = 'billingTitle'
        ws.cell(1,9).value = 'plancode'
        ws.cell(1,10).value = 'SumPremium'

        wb.save(Qfile)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")


    def do_lfg(self):

        Qfile = os.path.join(f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.year}\{self.year} Q{self.quarter} Letters\Lincoln Financial Group',f'{self.year} Q{self.quarter} LFG Billing-Inforce Files.xlsx')
        # Qfile = os.path.join(self.localdir,f'{self.year} Q{self.quarter} LFG Billing-Inforce Files.xlsx')
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")

        wb = xl.Workbook()

        ws1 = wb.active
        ws1.title = "Inforce"
        inforceqry = open(os.path.join(self.localdir,'Inforce-Billings Generator lfg inforce.sql')).read()
        data1, rowcount1, headers1 = self._getdata('pdxvmdb11','TAI_Feeds',inforceqry, year = self.year, quarter = self.Qquarter)
        for i in range(1,len(headers1)+1):
            ws1.cell(1,i).value = headers1[i-1]             
        self._pasteRange(1,2,194,1+rowcount1,ws1,data1.values.tolist())

        ws2 = wb.create_sheet("Billing")
        billingqry = open(os.path.join(self.localdir,'Inforce-Billings Generator lfg billing.sql')).read()
        data2, rowcount2, headers2 = self._getdata('pdxvmdb11','TAI_Feeds',billingqry, year = self.year, quarter = self.Qquarter)
        for i in range(1,len(headers2)+1):
            ws2.cell(1,i).value = headers2[i-1]             
        self._pasteRange(1,2,194,1+rowcount2,ws2,data2.values.tolist())

        wb.save(Qfile)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")

    def do_prudential(self):

        priorQfile = os.path.join(f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.prioryear}\{self.prioryear} Q{self.priorquarter} Letters\Prudential',f'{self.prioryear}Q{self.priorquarter} Premiums Paid to MU and OP.xlsx') 
        Qfile = os.path.join(f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.year}\{self.year} Q{self.quarter} Letters\Prudential',f'{self.year}Q{self.quarter} Premiums Paid to MU and OP.xlsx') 
        # Qfile = os.path.join(self.localdir,f'{self.year}Q{self.quarter} Premiums Paid to MU and OP.xlsx') 
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")

        # Roll forward previous quarter's file and rename
        if os.path.exists(Qfile)==False:
            shutil.copy(priorQfile,Qfile) 
            LOGGER.info(f'{os.path.basename(priorQfile)} is rolled forward as {os.path.basename(Qfile)}')

        # Create a new workbook
        wb = xl.load_workbook(Qfile)

        # Process Munich tab 
        mu = wb['Premiums to Munich']
        for row in mu['A3:GL5000']:
            for cell in row:
                cell.value = None        # Clean previous data by setting cell value to None
        mupxqry = open(os.path.join(self.localdir,'PRU Premiums Paid to MU.sql')).read()
        
        data1, rowcount1, headers1 = self._getdata('pdxvmdb11','TAI_Feeds',mupxqry,year = 2022,quarter = self.Qquarter)        
        self._pasteRange(1,3,194,2+rowcount1,mu,data1.values.tolist())
        mu.cell(1,2).value = f'Total {self.year} Q{self.quarter} Premium'
        for row in range(3, rowcount1+2+1):
            mu.cell(row,195).value = f'=VLOOKUP(TRIM(AN{row}), Treaties!$A$2:$C$26, 3, FALSE)'
            mu.cell(row,196).value = f'=IF(DZ{row}="-", -(DY{row}-EA{row}), (DY{row}-EA{row}))+IF(ED{row}="-", -(EC{row}-EE{row}), (EC{row}-EE{row}))+IF(EH{row}="-", -(EG{row}-EI{row}), (EG{row}-EI{row}))+IF(EL{row}="-", -(EK{row}-EM{row}), (EK{row}-EM{row}))+IF(EP{row}="-", -(EO{row}-EQ{row}), (EO{row}-EQ{row}))+IF(ET{row}="-", -(ES{row}-EU{row}), (ES{row}-EU{row}))'

        # Process Optimum data
        op = wb['Premiums to Optimum']
        for row in op['A3:GL5000']:
            for cell in row:
                cell.value = None
        oppyqry = open(os.path.join(self.localdir,'PRU Premiums Paid to OP.sql')).read()
        
        data2, rowcount2, headers2 = self._getdata('pdxvmdb11','TAI_Feeds',oppyqry, year = 2022, quarter = self.Qquarter)
        self._pasteRange(1,3,194,2+rowcount2,op,data2.values.tolist())
        op.cell(1,2).value = f'Total {self.year} Q{self.quarter} Premium'
        for row in range(3, rowcount2+2+1):
            op.cell(row,195).value = f'=VLOOKUP(TRIM(AN{row}), Treaties!$A$2:$C$26, 3, FALSE)'
            op.cell(row,196).value = f'=IF(DZ{row}="-", -(DY{row}-EA{row}), (DY{row}-EA{row}))+IF(ED{row}="-", -(EC{row}-EE{row}), (EC{row}-EE{row}))+IF(EH{row}="-", -(EG{row}-EI{row}), (EG{row}-EI{row}))+IF(EL{row}="-", -(EK{row}-EM{row}), (EK{row}-EM{row}))+IF(EP{row}="-", -(EO{row}-EQ{row}), (EO{row}-EQ{row}))+IF(ET{row}="-", -(ES{row}-EU{row}), (ES{row}-EU{row}))'

        # finish and save
        wb.save(Qfile)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")


    def do_rga(self):

        priorQfile = f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.prioryear}\{self.prioryear} Q{self.priorquarter} Letters\RGA\{self.prioryear} Q{self.priorquarter} RGA Claims Exhibit Breakout.xlsx'
        Qfile = f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{self.year}\{self.year} Q{self.quarter} Letters\RGA\{self.year} Q{self.quarter} RGA Claims Exhibit Breakout.xlsx'
        # Qfile = os.path.join(self.localdir,f'{self.year} Q{self.quarter} RGA Claims Exhibit Breakout.xlsx')
        LOGGER.info(f"{os.path.basename(Qfile)} process begin...")
        # Roll forward previous quarter's file
        if not os.path.exists(Qfile):
            shutil.copy(priorQfile,Qfile)
            LOGGER.info(f'{os.path.basename(Qfile)} is created at {os.path.dirname(Qfile)}')

        rgaqry = open(os.path.join(self.localdir,'Exhibit Support-RGA.sql')).read()
        data, rowcount, headers = self._getdata('PDXVMDEVOPS01','Claims',rgaqry, Year=self.year ,Quarter = self.quarter)

        wb = xl.load_workbook(Qfile)
        ws = wb.active

        for row in ws['A2:N5000']:
            for cell in row:
                cell.value = None

        self._pasteRange(1,2,14,1+data.shape[0],ws,data.values.tolist())

        wb.save(Qfile)
        LOGGER.info(f"{os.path.basename(Qfile)} process finished and saved at {os.path.dirname(Qfile)}.")


    def findzipfile(self): 

        ziplist = {
        'Ace':['TAIQ51AT.Txt','TAIQ52AT.Rep','TAIQ52AT.Txt','TAIQ53AT.Rep','TAIQ53AT.Txt'],
        'CanadaLife':['TAIQ51CL.Txt','TAIQ52CL.Rep','TAIQ52CL.Txt','TAIQ53CL.Rep','TAIQ53CL.Txt',
                        'TAIQPCCL.Rep','TAIQPCCL.Txt','TAIQPECL.Rep','TAIQPECL.Txt','TAIQPTCL.Rep',
                        'TAIQPTCL.Txt','TAIQSPCL.Rep','TAIQSPCL.Txt','TAIQTPCL.Rep','TAIQTPCL.Txt'],
        'BH':['TAIQ51BH.Txt','TAIQ52BH.Rep','TAIQ52BH.Txt','TAIQ53BH.Rep','TAIQ53BH.Txt'],
        'Generali':['TAIQ51GL.Txt','TAIQ52GL.Rep','TAIQ52GL.Txt','TAIQ53GL.Rep','TAIQ53GL.Txt'],
        'Hannover':['TAIQ51HL.Txt','TAIQ52HL.Rep','TAIQ52HL.Txt','TAIQ53HL.Rep','TAIQ53HL.Txt'],
        'ING':['TAIQ51IN.Txt','TAIQ52IN.Rep','TAIQ52IN.Txt','TAIQ53IN.Rep','TAIQ53IN.Txt'
                    ,'TAIQSPIN.Rep','TAIQSPIN.Txt','TAIQTPIN.Rep','TAIQTPIN.Txt'],
        'JohnHancock':['TAIQ51JH.Txt','TAIQ53JH.Rep','TAIQ53JH.Txt','TAIQSPJH.Rep','TAIQSPJH.Txt','TAIQTPJH.Rep','TAIQTPJH.Txt'],
        'Lincoln':['TAIQ51LN.Txt','TAIQ52LN.Rep','TAIQ52LN.Txt','TAIQ53LN.Rep','TAIQ53LN.Txt',
                    'TAIQSPLN.Rep','TAIQSPLN.Txt','TAIQTPLN.Rep','TAIQTPLN.Txt'],
        'LFG':['TAIQ51LF.Txt','TAIQ52LF.Rep','TAIQ52LF.Txt','TAIQ53LF.Rep','TAIQ53LF.Txt',
                'TAIQSPLF.Rep','TAIQSPLF.Txt','TAIQTPLF.Rep','TAIQTPLF.Txt'],
        'Munich':['TAIQ51MU.Txt', 'TAIQ52MU.Rep', 'TAIQ52MU.Txt', 'TAIQ53MU.Rep', 'TAIQ53MU.Txt', 
                    'TAIQSPMU.Rep', 'TAIQSPMU.Txt', 'TAIQTPMU.Rep', 'TAIQTPMU.Txt'],
        'Optimum':['TAIQ51OP.Txt', 'TAIQ52OP.Rep', 'TAIQ52OP.Txt', 'TAIQ53OP.Rep', 'TAIQ53OP.Txt'],
        'Aurigen':['TAIQ51AU.Txt', 'TAIQ52AU.Rep', 'TAIQ52AU.Txt', 'TAIQ53AU.Rep', 'TAIQ53AU.Txt', 
                    'TAIQPTAU.Rep', 'TAIQPTAU.Txt', 'TAIQSPAU.Rep', 'TAIQSPAU.Txt'],
        'Pru':['TAIQPCMU.Rep', 'TAIQPCMU.Txt', 'TAIQPCOP.Rep', 'TAIQPCOP.Txt', 'TAIQPCPR.Rep', 
                'TAIQPCPX.Rep', 'TAIQPCPX.Txt', 'TAIQPCPY.Txt', 'TAIQPEMU.Rep', 'TAIQPEMU.Txt', 
                'TAIQPEOP.Rep', 'TAIQPEOP.Txt', 'TAIQPEPX.Rep', 'TAIQPEPX.Txt', 'TAIQPEPY.Rep', 
                'TAIQPEPY.Txt', 'TAIQPTMU.Rep', 'TAIQPTMU.Txt', 'TAIQPTOP.Rep', 'TAIQPTOP.Txt', 
                'TAIQPTPX.Rep', 'TAIQPTPX.Txt', 'TAIQPTPY.Rep', 'TAIQPTPY.Txt'],
        'RGA':['TAIQ51RG.Txt', 'TAIQ52RG.Rep', 'TAIQ52RG.Txt', 'TAIQ53RG.Rep', 'TAIQ53RG.Txt'],
        'Scor':['TAIQ51FR.Txt', 'TAIQ52FR.Rep', 'TAIQ52FR.Txt', 'TAIQ53FR.Rep', 'TAIQ53FR.Txt'],
        'Standard':['TAIQ51SB.Txt', 'TAIQ52SB.Rep', 'TAIQ52SB.Txt', 'TAIQ53SB.Rep', 'TAIQ53SB.Txt'],
        'Swiss':['TAIQ51SW.Txt', 'TAIQ52SW.Rep', 'TAIQ52SW.Txt', 'TAIQ53SW.Rep', 'TAIQ53SW.Txt']}

        firstsrc = r'\\pdxvmdbtai01\TAIProdReports'
        secondsrc = self.localdir
        # zipdst = f'J:\MLife\Reinsurance\Quarterly Financial Reports\Quarterly Letters to Carriers\{year}\{year} Q{quarter} Letters\ACE\Ace'

        for carrier in ziplist:
            zipdst = os.path.join(os.getcwd(),str(carrier))
            if not os.path.exists(zipdst): 
                os.mkdir(zipdst)
                LOGGER.debug(f'{zipdst} is created.')
    
            for i in ziplist[carrier]:
                LOGGER.debug(f'searching {i} for carrier {carrier}')
                if not os.path.exists(os.path.join(zipdst,i)):
                    try:
                        shutil.copy(os.path.join(firstsrc,i),zipdst)
                        LOGGER.info(f'{carrier} - {i} is copied to {zipdst}')
                    except FileNotFoundError:
                        LOGGER.info(f'Cannot find {i} in {firstsrc}')
                        LOGGER.debug(f'Looking for {i} in {secondsrc}')
                        try:
                            shutil.copy(os.path.join(secondsrc,i),zipdst)
                            LOGGER.info(f'{carrier} - {i} is copied to {zipdst}')
                        except FileNotFoundError:
                            LOGGER.info(f'{carrier} - Cannot find {i} in either {firstsrc} or {secondsrc}.')            
                else:
                    LOGGER.info(f'{carrier} - {i} has existed in {zipdst}')




    def _getdata(self, server, db, qry, **kwargs):        
        '''
        Take server name, database name and query as inputs and return the pandas dataframe. 
        Set up a list of argument for the query if needed. 
        Parameters are fed into "?" in a query by order in the list.
        '''
        parameters=[]
        if kwargs:
            for key, value in kwargs.items():
                parameters.append(value)
            LOGGER.debug(f'Use parameters {parameters} for the query...')
        else:
            parameters = None

        # Connecting to the database using sqlalchemy
        LOGGER.debug('Connecting to {} {}...'.format(server,db))        
        engine = create_engine(f"mssql+pyodbc://@{server}/{db}?driver=ODBC+Driver+17+for+SQL+Server")
        conn = engine.connect()

        # return the dataframe with the connection along with rowcount and headers, then close the connection
        df = pd.read_sql(qry, conn, params=parameters)
        rowcount = df.shape[0]
        headers = df.columns.to_list()
        LOGGER.debug(f'Retrieve {rowcount} rows from the sql query...')
        conn.close()
        LOGGER.debug('Connection closed. Processing the workbook for carrier...')
        
        return df, rowcount, headers


    def _pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving ,copiedData):
        """
        Paste a list to a specific range of an excel worksheet. 
        Note that copiedData has to be a list! 
        Use df.values.tolist() method to convert dataframe to a list if needed.
        """
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1


    def filterJH(self):
        ### The following is for TAIQ52INJH.txt TAIQ52LNJH.txt TAIQ52RGJH.txt TAIQ52SWJH.txt
        ### 1.copy these four files from \\pdxvmdbtai01\TAIProdReports to J:\MLife\Reinsurance\Carriers\John Hancock\3 Party Reinsurance Files\YYYY QQ: 
        ### TAIQ52INJH.txt TAIQ52LNJH.txt TAIQ52RGJH.txt TAIQ52SWJH.txt
        ### 2.run the script below
        
        filesource = r'\\pdxvmdbtai01\TAIProdReports'
        search = ['TAIQ52IN.Txt','TAIQ52LN.Txt','TAIQ52RG.Txt','TAIQ52SW.Txt']
        filepath = os.path.join(r'J:/MLife/Reinsurance/Carriers/John Hancock/3 Party Reinsurance Files/',str(self.year)+' Q'+str(self.quarter))

        for f in os.listdir(filesource):
            if f in search:
                shutil.copy2(os.path.join(filesource,f),os.path.join(filepath,f))
                LOGGER.info(f'{f} is sent to {filepath}')
        
        cedingCoPosition = 0
        treatyIDPosition = 139


        cedingCo = []
        cedingCo = ['1ML','1MW','1NY','1PM','1SL','2JH','LFG','LNL','M84', 'NWL', 'PRU','SNL','TCL', 'SYM']
        treatyID = []
        treatyID = ['WLMAJSUA','OPMAJUNA','HLPRFUWA','RGFCJHES','MUPRFSVB','OPPRFUWB','MQMBJGIA','GLMBJGIA','MRMAJUNB','MRMAJULB','OPMBGFAC','MQMBGGIB','ATMAJSUA','MUPRFUWC','CLMBGSVB','OPPRFUUD','MUPRFSVD','MUPRFUWB','OPMAJSVA','MRMAJULA','OPPRFUWA','RGMBGUWA','RGPRFUUA','HLPRFUUD','CQMBGGIA','MRMAJSVC','MRMBJGIA','RGPRFUWA','RGMBGSVB','RQMBGGIB','CLMBGUWB','OPJHFAC ','HLMBGSVB','MUJHFYRT','MRMAJSVA','MUMBGUWA','MUPRFSVC','MRVCOLGC','OPPRFUWC','MRMAJSUD','HLPRFSVD','CLMGFCSV','OPPRFUUC','OPPRFSVA','JHNNLN3A','RGFCJHER','WLMAJS-A','MUMBGGIB','WLMAJUNA','MRMAJULC','HLMGFCSV','MUPRFUWD','OPMAJSUA','MQMAGGIA','RGMGFCSV','GLPREUWA','MUMBJGIA','ATMAJULA','RGMBGGIB','MRMAJSVD','MRMAJGIA','HLPRFSVC','WLMAJS+A','CLMBGUWA','GLFCJHSV','RGPRFUWB','OPPRFSVB','GLPRFULB','OPMBGSVB','HLMBGFAC','OPFCJHSV','HLPRFUWB','MRMAJSVB','RQMBGGIA','GLPRFULA','HLPRFSVA','RGPREUWA','MUMBGFAC','OPPRFUUA','OPPRFSVC','WLMAJULA','MRVCOLGA','RGPRFUUD','RGPRFUWC','GLPRFSVA','OPMAJULA','OPJUMFAC','HQMBGGIA','OPPREUWA','MUMBGSVB','HLMBJGIA','TAMAGFUA','ATMAJUNA','GLPRFUUA','HLMBGUWA','MRVCOLIA','HLFACJH ','MRMAJUNA','RGPRFSVB','OPMBGUWB','GLFACJHE','RGMBGUWB','MUPREUWA','OPPRFSVD','MUMBGFC2','MUPRFUUD','MQMBGGIA','GLPRFSVC','HLPRFUWD','RQMBJGIA','GLPRFULC','MRMAJGIB','HLMBGGIB','OPMGFCSV','HQMBGGIB','MQMGFCGI','MRVCOLIC','JHNNLN4A','GQMBJGIA','MUMBGUWB','MRMAJGIC','HLPRFUWC','GLPRFUUC','HLPRFSVB','MRMAJS+D','MUFACJH2','HLPRFUUA','GLPRFSVB','RGPRFSVD','HQMAGGIA','RGPRFUUC','MRVCOLGB','MRMAJS-D','MUPRFUUC','RQPLEGIA','OPPRFUWD','ATMAJSVA','MUFACJH ','RGMBJGIA','MRMAJUNC','HLPRFUUC','RGPRFSVC','HLPREUWA','HLMBGUWB','RGPRFSVA','MUPRFUWA','OPJMFCSV','MRVCOLIB','HQMBJGIA','RGFACJH ','HXPMSV5A','HXPMUU5A','HXPMUW5A','MUPRFUUA','RGPRFUWD','OPMBGUWA','HQMBJGIB','HXPMQS1A','MQMBJGIB','RQMBJGIB','CQMBGGIA','MUPRFSVA','HLMBGUWC','HXPMUW3A','MUMBGUWC','OPMBGUWC','RGMBGUWC','HLFACJH1','MUFACJH1','MUFACJH3','OPFCJH1A','HLMBGSVC','HXPMSV3A','MUMBGSVC','OPMBGSVC','RGMBGSVC','HLJHXSJA','HLJHXSSA','MUJHXSJA','MUJHXSSA','OPJHXSJA','OPJHXSSA','RGJHXSJA','RGJHXSSA','RGMBGFAC','OPJHCEJB','OPJHCEJC','OPJHCIEA','RGPRFUWE','INMAGF+A','INMAGF-A','INMAGFUA','INMAGGIC','INMAGS+C','INMAGS-C','INMAGUNC','IQMAGGIC']


        for file in os.listdir(filepath):
            if file in search:
                sourceFile = open(os.path.join(filepath,file), 'r')
                modFile = open(filepath + '/' + file[:8] + 'JH.txt', 'w')
                for line in sourceFile:
                    if line[cedingCoPosition:3] not in cedingCo:
                        if line[treatyIDPosition:treatyIDPosition+8] not in treatyID:
                            modFile.write(line)
                sourceFile.close()
                modFile.close()
                LOGGER.info(filepath + '/' + file[:8] + 'JH.txt is successfully created.')





le = letters(2022, 4)
# le.do_ace()
# le.do_hannover()
# le.do_jh()
# le.do_lfg()
# le.do_prudential()
# le.do_rga()
# le.findzipfile()
# le.filterJH()






