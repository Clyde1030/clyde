from datetime import date
import openpyxl as xl
import xlwings as xw
from xlwings.utils import rgb_to_int
from calendar import month_name
import shutil
import pandas as pd
import logging
import re
import os

LOGGER = logging.getLogger(__name__)
# LOGGER = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
# handler = logging.StreamHandler()
# handler.setFormatter(formatter)
# LOGGER.addHandler(handler)
# LOGGER.setLevel(logging.DEBUG)

class pennmutual:

    def __init__(self, year, month, *csvdst):
        self.year = year
        self.month = month
        self.monthApplied = str(month).zfill(2)
        self.carrierId = 1690
        self.regex = re.compile('PML_MGroup_CompFeed_.*')
        self.csvname = f'P-1690-LIF-{year}-{self.monthApplied}-1.txt'
        self.csvdst = csvdst or f'C:\dev\Production\data\{year}\{self.monthApplied}'
        self.datadir = f'J:\Acctng\Production\{year}\Data\Penn Mutual'
        self.flash = os.path.join(f'C:\dev\Production',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        # self.flash = os.path.join(f'J:\Acctng\Production\{year}\Summary - Flash',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        self.exportCols = ["SourceFileName","CarrierID","MemFirmID","CarrierContracteeID","CarrierContracteeName",
                            "ProductID","CarrierProductID","YearApplied","MonthApplied","ProducerID",
                            "CarrierProducerID","CarrierProducerName","PolicyNumber","IssueDate","InsuredName",
                            "YTDAnnualizedPrem" ,"YTDAnnualizedLowNon","YTDFace","RiskNumber","RiskName",
                            "Exchange1035","SplitPercentage","Replacement","CarrierProductName","PolicyOwner",
                            "Exchange","NLG","Modco","YRT","CSO2001"]

    def _copyRange(self,startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        for i in range(startRow,endRow + 1,1):
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            rangeSelected.append(rowSelected)
        return rangeSelected
                
    def _pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving ,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def getrawmonth(self, year=None, month=None):
        '''Search and move override data to the data directory for a single month'''
        
        isFileFound = True

        if not year:
            year = self.year
        if not month:
            month = self.month

        LOGGER.info(f'Searching {month_name[self.month]} file for carrierid {self.carrierId}...')
        path = []
        f = []
        for dirs, _, filenames in os.walk(f'J:\Acctng\Revenue\REVENUE\Penn Mutual\{year}\{month_name[month]}'):
            f = [os.path.join(dirs,f) for f in filenames if re.search(self.regex,os.path.join(dirs,f))]
            path.extend(f)
        if len(path) == 0:
            LOGGER.warning(f'1690 Penn Mutual {month_name[self.month]} raw files not available.')
        else:
            for p in path:
                if not os.path.exists(os.path.join(self.datadir,str(month).zfill(2),os.path.basename(p))):
                    shutil.copy2(p,os.path.join(self.datadir, str(month).zfill(2)))
                    LOGGER.info(f'{os.path.basename(p)} is copied to the data directory.')
                else:
                    LOGGER.info(f'{os.path.basename(p)} is already in the data directory.')
        return isFileFound


    def _ytdlist(self, year=None, month=None):
        '''Walks the data directory. Determine what override files should be in the YTD list as of the month. Return the file paths'''
        if not year:
            year = self.year
        if not month:
            month = self.month
        ytdmonth = []
        path = []
        for i in range(1,month+1):
            if i <= month:
                ytdmonth.append(str(i).zfill(2))
        for m in ytdmonth:
            for dirs, _, filenames in os.walk(os.path.join(f'J:\Acctng\Production\{year}\Data\Penn Mutual',m)): 
                files = [os.path.join(dirs,f) for f in filenames if re.search(self.regex,os.path.join(dirs,f))]
                path.extend(files)
        LOGGER.debug(f'1690 Penn Mutual: {len(path)} override files are appended to the list YTD {year} {month_name[month]}.')
        return path

    def aggregateYTD(self, year=None, month=None): 
        '''Search each month's folder from January to the processing month in the data directory. Make a list of these override files 
        and aggregate them all into one YTD file'''

        if not year:
            year = self.year
        if not month:
            month = self.month

        # Create a file list that are override files and needed to be aggregated
        overrides = self._ytdlist(year, month)
        LOGGER.debug(f'Start combining {len(overrides)} files in the list...')
        MFull = xl.Workbook()
        full = MFull.active
        headers = ["Report Date","Override Date","REC_TYPE","AGENT_NUMBER","AGENT_NAME",
        	        "PML_AGENCY_NUMBER","PML_AGENCY_NAME","POLICY_NUMBER","CYCLE_DATE","PAYMENT_MODE",
                    "DURATION","ISSUE_DATE","TRANSACTION_PAID_DATE","INSURED_NAME","POLICY_OWNER",
                    "PRODUCT_NAME","PRODUCT_CODE","ISSUE_STATE","COMMISSION_OPTION","REVENUE_SUBTYPE",
                    "TARGET_PREMIUM","PAID_EXCESS_PREMIUM","PRODUCER_SPLIT_PERCENTAGE","COMP_TYPE","PRODUCER_COMP_RATE",
                    "PRODUCER_COMP_AMOUNT","M_OVERRIDE_RATE","M_OVERRIDE_TOTAL_AMOUNT","OFFC_CD","PRDCR_UNDRWRITR_NMBR"]
    
        recordCounts = [0] # contains the row counts where each override worksheet will be copy from
        # Copy over headers from the first sheet. From the second sheet and on, the first row will not be copied 
        for i, inProgress in enumerate(overrides, start = 0):
            LOGGER.info(f'now processing {inProgress}...')
            # Report month for the first column
            reportmonth = inProgress.split('\\')[6][0:2]
            # Override date info for the second column
            filename = os.path.basename(inProgress)
            overridemonth = filename.split('_')[3][4:6]
            overrideday = filename.split('_')[3][6:8]
            overridedate = date(year,int(overridemonth),int(overrideday)).strftime("%m/%d/%Y")

            processingSheet = xl.load_workbook(inProgress).worksheets[0]
            rowcount = len(processingSheet['A'])-1

            datapCopying = self._copyRange(startCol=1,startRow=2, endCol = 28, endRow=rowcount+1,sheet=processingSheet)
            self._pasteRange(startCol=3,startRow=1+sum(recordCounts),endCol=30,endRow=rowcount+sum(recordCounts),sheetReceiving=full,copiedData=datapCopying)
            for row in range(sum(recordCounts)+1,sum(recordCounts)+rowcount+1,1):
                full.cell(row,2).value = overridedate
                full.cell(row,1).value = year * 100 + int(reportmonth)
            recordCounts.append(rowcount)
        
        full.insert_rows(1)
        for i in range(len(headers)):
            full.cell(1,i+1).value = headers[i]         
        # # Apply filter and save the completed workbook
        full.auto_filter.ref = f'A1:AD{sum(recordCounts)}' 
        MFull.save(os.path.join(self.datadir,str(month).zfill(2),f'PennMutual{year*100+month}YTD.xlsx'))
        LOGGER.info(f'1690 Penn Mutual: PennMutual{year*100+month}YTD.xlsx generated.')

    def getcsvYTDv1(self):

        if not year:
            year = self.year
        if not month:
            month = self.month

        LOGGER.info(f'Processing Penn Mutual with {year} {month_name[month]} YTD data...')

        df = pd.read_excel(os.path.join(self.datadir,str(month).zfill(2),f'PennMutual{year*100+month}YTD.xlsx'),header=0,converters={'POLICY_NUMBER':str},engine='openpyxl')
        df.iloc[:,23] = df.iloc[:,23].str.rstrip()

        # Filter out "M FINANCIAL HOLDINGS INC" and "M HOLDINGS SECURITIES INC" from full["Agency Name"] and full["M Company Receiving Payment"]
        df = df[df.iloc[:,2] != 'BD Record']
        df = df[(df.iloc[:,23] == 'First Year Commission')] # filter column X to keep only "First Year Commission"
        df = df[~df.iloc[:,25].isin([0])]

        df.to_excel(os.path.join(self.datadir,f'PennMutual{year*100+month}YTDfiltered.xlsx'),index=0)
        LOGGER.info(f'1690 Penn Mutual: PennMutual{year*100+month}YTD.xlsx filtered.')

        df["SourceFileName"] = f"P-1690-LIF-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = ""
        df["CarrierContracteeName"] = ""
        df["ProductID"] = 0
        df["CarrierProductID"] = df['PRODUCT_CODE'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['AGENT_NUMBER'].str.rstrip()
        df["CarrierProducerName"] = df['AGENT_NAME'].str.rstrip()
        df["PolicyNumber"] = df['POLICY_NUMBER'].str.lstrip()
        df["IssueDate"] = df['ISSUE_DATE']
        df["InsuredName"] = df['INSURED_NAME'].str.rstrip()
        df["PRODUCER_COMP_RATE_Mod"] = df.apply(lambda x: x["PRODUCER_COMP_RATE"] if pd.isnull(x["PRODUCER_COMP_RATE"])==False 
                                                    else (0.025 if x['REVENUE_SUBTYPE']=='OVER TARGET'
                                                            else 0.5 if (x['REVENUE_SUBTYPE'] == 'UP TO TARGET' or x['REVENUE_SUBTYPE'] == 'NON-TARGETED') 
                                                                    else x["PRODUCER_COMP_RATE"]),axis=1)
        df["YTDAnnualizedPrem"] = df.apply(lambda x: 0 if x['REVENUE_SUBTYPE']=='OVER TARGET' else x['PRODUCER_COMP_AMOUNT']/x['PRODUCER_COMP_RATE_Mod'], axis = 1) #.astype(float).round(2).map("{:,.2f}".format) 
        df["YTDAnnualizedLowNon"] = df.apply(lambda x: x['PRODUCER_COMP_AMOUNT']/x['PRODUCER_COMP_RATE_Mod'] if x['REVENUE_SUBTYPE']=='OVER TARGET' else 0, axis = 1) #.astype(float).round(2).map("{:,.2f}".format)       
        df["YTDFace"] = ""
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = ""
        df["SplitPercentage"] = df["PRODUCER_SPLIT_PERCENTAGE"]
        df["Replacement"] = ""
        df["CarrierProductName"] = df['PRODUCT_NAME'].str.rstrip()
        df["PolicyOwner"] = df['POLICY_OWNER'].str.rstrip()
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""    

        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD = {excess}')
        
        LOGGER.info(f'Accessing flash workbook and updating values...')        
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('PennM_target').value = target
            notesbreakdown.range('PennM_excess').value = excess
            notesbreakdown.range('PennM_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('PennM_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df['Target_Premium'].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df['Excess_Premium'].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst, 'test.xlsx'), index = False)
        LOGGER.info(f'{self.csvname} is saved at {self.csvdst}.')

    def getcsvYTDv2(self, year=None, month=None):

        if not year:
            year = self.year
        if not month:
            month = self.month

        LOGGER.info(f'Processing Penn Mutual with {year} {month_name[month]} data...')
        df = pd.read_excel(os.path.join(self.datadir,str(month).zfill(2),f'MFin_YTD_Trans_{month_name[month][0:3]}{year-2000}.xlsx'),header=0,converters={'POLICYACCOUNT_NMBR':str}, engine='openpyxl')
        df = df[pd.isnull(df['POLICYACCOUNT_NMBR'])==False]


        df["SourceFileName"] = f"P-1690-LIF-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = ""
        df["CarrierContracteeName"] = ""
        df["ProductID"] = 0
        df["CarrierProductID"] = df['PROD_CD'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['UW_NUMBER'].astype(str)
        df["CarrierProducerName"] = df['AGENT'].str.rstrip()
        df["PolicyNumber"] = df['POLICYACCOUNT_NMBR'].str.lstrip()
        df["IssueDate"] = ""
        df["InsuredName"] = df['INSRD_LAST_NM'].str.rstrip()
        df["YTDAnnualizedPrem"] = df.apply(lambda x: 0 if x['RVNUE_SUBTYPE_NM']=='OVER TARGET' else x['PAID_PREMIUM'], axis = 1) #.astype(float).round(2).map("{:,.2f}".format) 
        df["YTDAnnualizedLowNon"] = df.apply(lambda x: x['PAID_PREMIUM'] if x['RVNUE_SUBTYPE_NM']=='OVER TARGET' else 0, axis = 1) #.astype(float).round(2).map("{:,.2f}".format)       
        df["YTDFace"] = df["FACE SOLD"]
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = ""
        df["SplitPercentage"] = df["POL_SOLD"]
        df["Replacement"] = ""
        df["CarrierProductName"] = df['PROD_CD'].str.rstrip()
        df["PolicyOwner"] = df['INSRD_LAST_NM'].str.rstrip()
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""    

        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD = {excess}')
        
        LOGGER.info(f'Accessing flash workbook and updating values...')        
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('PennM_target').value = target
            notesbreakdown.range('PennM_excess').value = excess
            notesbreakdown.range('PennM_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('PennM_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df['YTDAnnualizedPrem'].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df['YTDAnnualizedLowNon'].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst, 'test.xlsx'), index = False)
        LOGGER.info(f'{self.csvname} is saved at {self.csvdst}.')


def processPenn(year, month):
    penn = pennmutual(year,month)
    
    try:
        penn.getcsvYTDv2(year, month)
    except FileNotFoundError:
        LOGGER.warning(f'Penn Mutual {month_name[month]} data not available.')
        try: 
            if month == 1:
                penn.getcsvYTDv2(year-1, 12)
            else:
                penn.getcsvYTDv2(year,month-1)
        except FileNotFoundError:
            LOGGER.warning(f'Penn Mutual {month_name[month-1]} data not available.')


# if __name__ == '__main__':
#     penn = pennmutual(2022,12)    
#     penn.getrawmonth(2022,12)
#     penn.aggregateYTD()
#     penn.getcsvYTDv1()















# Archive
#############################################################################################################################################################################
# Function merges a single month's overrides files. Replace with aggreagteYTD 
    # def aggregate(self):
    #     
        
    #     # Create a file list that are override files and needed to be aggregated
    #     overrides = [os.path.join(self.datadir, f) for f in os.listdir(self.datadir) if re.search(self.regex, os.path.join(self.datadir, f))] 
        
    #     LOGGER.debug(f'Start combining {len(overrides)} files in {self.datadir}...')
    #     MFull = xl.Workbook()
    #     full = MFull.active
    #     recordCounts = [] # contains the row counts where each override worksheet will be copy from
        
    #     # Copy over headers from the first sheet. From the second sheet and on, the first row will not be copied 
    #     i = 0
    #     while i < len(overrides):
    #         processingSheet = xl.load_workbook(overrides[i]).worksheets[0]
    #         inProgress = overrides[i]
    #         filename = os.path.basename(inProgress)
    #         fileMonth = filename.split('_')[3][4:6]
    #         fileDay = filename.split('_')[3][6:8]
    #         fileDate = date(self.year,int(fileMonth),int(fileDay)).strftime("%m/%d/%Y")
    #         rowcount = len(processingSheet['A'])
    #         if i == 0:
    #             recordCounts.append(rowcount)
    #             datapCopying = self._copyRange(startCol=1,startRow=1, endCol = 28, endRow=rowcount,sheet=processingSheet)
    #             self._pasteRange(startCol=3,startRow=1,endCol=30,endRow=rowcount,sheetReceiving=full,copiedData=datapCopying)
    #             for row in range(2,len(processingSheet['A'])+1,1):
    #                 full.cell(row,2).value = fileDate
    #         else:
    #             recordCounts.append(rowcount-1) 
    #             datapCopying = self._copyRange(startCol=1,startRow=2, endCol = 28, endRow=rowcount,sheet=processingSheet)
    #             self._pasteRange(startCol=3,startRow=len(full['C'])+1,endCol=30,endRow=len(full['C'])+rowcount-1,sheetReceiving=full,copiedData=datapCopying)
    #             for row in range(sum(row for row in recordCounts)-recordCounts[i]+1,sum(row for row in recordCounts)+1,1):    
    #                 full.cell(row,2).value = fileDate
    #         i += 1
        
    #     # Fill the first column "date"
    #     lastrowfull = sum(row for row in recordCounts)
    #     full.cell(1,1).value = 'ReportDate'
    #     full.cell(1,2).value = 'OverrideDate'        
    #     for r in range(2,lastrowfull+1,1):
    #         full.cell(r,1).value = self.year * 100 + self.month
        
    #     # Apply filter and save the completed workbook
    #     full.auto_filter.ref = f'A1:AD{lastrowfull}' 
    #     MFull.save(os.path.join(self.datadir,f'{month_name[self.month]}Full.xlsx'))
    #     LOGGER.info(f'1690 Penn Mutual: {month_name[self.month]}Full.xlsx generated.')




















    

