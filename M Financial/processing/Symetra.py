from datetime import date
import openpyxl as xl
from xlwings.utils import rgb_to_int 
import xlwings as xw
from calendar import month_name
import shutil
import pandas as pd
import logging
import re
import os

LOGGER = logging.getLogger(__name__)
# LOGGER = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
# handler = logging.StreamHandler()
# handler.setFormatter(formatter)
# LOGGER.addHandler(handler)
# LOGGER.setLevel(logging.DEBUG)

class symetra:
    
    def __init__(self, year, month, *csvdst):
        self.year = year
        self.month = month
        self.monthApplied = str(month).zfill(2)
        self.carrierId = 1680
        self.regex = re.compile(".Reports_Commission_MHoldings_wkof_") 
        self.datadir = f'J:\Acctng\Production\{self.year}\Data\Symetra'
        self.csvname = f'P-1680-LIF-{year}-{self.monthApplied}-1.txt'
        self.csvdst = csvdst or f'C:\dev\Production\data\{year}\{self.monthApplied}'
        self.flash = os.path.join(f'C:\dev\Production',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        # self.flash = os.path.join(f'J:\Acctng\Production\{year}\Summary - Flash',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        self.exportCols = ["SourceFileName","CarrierID","MemFirmID","CarrierContracteeID","CarrierContracteeName",
                            "ProductID","CarrierProductID","YearApplied","MonthApplied","ProducerID",
                            "CarrierProducerID","CarrierProducerName","PolicyNumber","IssueDate","InsuredName",
                            "YTDAnnualizedPrem" ,"YTDAnnualizedLowNon","YTDFace","RiskNumber","RiskName",
                            "Exchange1035","SplitPercentage","Replacement","CarrierProductName","PolicyOwner",
                            "Exchange","NLG","Modco","YRT","CSO2001"]

    def _copyRange(self,startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected

    def _pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1


    def getrawmonth(self, year=None, month=None):
        '''Symetra's override files come from Peggy every month. This fuction fetches the corresponding month's override data from the override directory. 
        Parameters year and month is default to self.year and self.month if not specified'''

        isFileFound = True

        if not year:
            year = self.year
        if not month:
            month = self.month

        LOGGER.info(f'Searching {month_name[month]} file for carrierid {self.carrierId}...')
        path = []
        f = []
        for dirs, _, filenames in os.walk(os.path.join(f'J:\Acctng\Revenue\REVENUE\Symetra\{year} Overrides\{month_name[month]}')): 
            f = [os.path.join(dirs,f) for f in filenames if re.search(self.regex,os.path.join(dirs,f))]
            path.extend(f)
        if len(path) == 0:
            LOGGER.warning(f'1680 Symetra {month_name[self.month]} raw files not available.')
        else:
            for p in path:
                if not os.path.exists(os.path.join(self.datadir,str(month).zfill(2),os.path.basename(p))):
                    shutil.copy2(p,os.path.join(self.datadir, str(month).zfill(2)))
                    LOGGER.info(f'{os.path.basename(p)} is copied to the data directory.')
                else:
                    LOGGER.info(f'{os.path.basename(p)} is already in the data directory.')
        return isFileFound

    def _ytdlist(self, year=None, month=None):
        '''Walks the data directory. Determine what override files should be in the YTD list as of the month. Return the file paths'''
        if not year:
            year = self.year
        if not month:
            month = self.month
        ytdmonth = []
        path = []
        for i in range(1,month+1):
            if i <= month:
                ytdmonth.append(str(i).zfill(2))
        for m in ytdmonth:
            for dirs, _, filenames in os.walk(os.path.join(f'J:\Acctng\Production\{year}\Data\Symetra',m)): 
                files = [os.path.join(dirs,f) for f in filenames if re.search(self.regex,os.path.join(dirs,f))]
                path.extend(files)
        LOGGER.debug(f'1680 Symetra: {len(path)} override files are appended to the list YTD {year} {month_name[month]}.')
        return path

    def aggregateYTD(self, year=None, month=None):
        '''Search each month's folder from January to the processing month in the data directory. Make a list of these override files 
        and aggregate them all into one YTD file'''
        
        if not year:
            year = self.year
        if not month:
            month = self.month

        # Create a file list that are override files and needed to be aggregated
        overrides = self._ytdlist(year,month)
        # overrides = [os.path.join(self.datadir,str(month).zfill(2),f) for f in os.listdir(os.path.join(self.datadir,str(month).zfill(2))) if re.search(self.regex, os.path.join(self.datadir,str(month).zfill(2))f)] 
        LOGGER.debug(f'Combining {len(overrides)} files in the list...')
        MFull = xl.Workbook()
        full = MFull.active
        recordCounts = [0] # contains the row counts where each override worksheet will be copy from
        headers = ["Report Date","Override Date","Agent Number","Agent Name","Agency Number",
                   "Agency Name","Policy Number","Cycle Date","Mode","Duration","Issue Date",
                   "Transacation Date","Insured Name","Product Name","Product Code","Paid Target Premium", 	 
                   "Paid Excess Premium","Asset Basis","Producer Policy Split Percentage","Producer Commission Rate",
                   "Producer Commission Amount","M Company Receiving Payment","M Override Rate","M Override Total Amount","Transaction Type","Trans Desc"]
        # Copy over headers from the first sheet. From the second sheet and on, the first row will not be copied 
        for i, inProgress in enumerate(overrides, start = 0):
            LOGGER.info(f'now processing {inProgress}...')
            filename = os.path.basename(inProgress)
            reportmonth = inProgress.split('\\')[6][0:2]
            overridemonth = inProgress.split('_')[4][0:2]
            overrideday = filename.split('_')[4][2:4]
            overridedate = date(year,int(overridemonth),int(overrideday)).strftime("%m/%d/%Y")
            sheet = xl.load_workbook(overrides[i]).worksheets[1]
            rowcount = len(sheet['A'])-1
            datapCopying = self._copyRange(startCol=1,startRow=2, endCol = 24, endRow=1+rowcount,sheet=sheet)
            self._pasteRange(startCol=3,startRow=sum(recordCounts)+1,endCol=26,endRow=sum(recordCounts)+rowcount,sheetReceiving=full,copiedData=datapCopying)
            for row in range(sum(recordCounts)+1,sum(recordCounts)+rowcount+1,1):
                full.cell(row,2).value = overridedate
                full.cell(row,1).value = year * 100 + int(reportmonth)
            recordCounts.append(rowcount)
        full.insert_rows(1)
        for i in range(len(headers)):
            full.cell(1,i+1).value = headers[i]         
        # # Apply filter and save the completed workbook
        full.auto_filter.ref = f'A1:Z{sum(recordCounts)+1}' 
        MFull.save(os.path.join(self.datadir,str(month).zfill(2),f'Symetra{year*100+month}YTD.xlsx'))
        LOGGER.info(f'1680 Symetra: Symetra{year*100+month}YTD.xlsx generated.')


    def getcsv(self, year=None, month=None):


        if not year:
            year = self.year
        if not month:
            month = self.month

        LOGGER.info(f'Processing Symetra with {year} {month_name[month]} YTD data...')

        df = pd.read_excel(os.path.join(self.datadir,str(month).zfill(2),f'Symetra{year*100+month}YTD.xlsx'),header=0,
            converters={"Agency Number": str,"Policy Number":str, "Issue Date":str,"Producer Policy Split Percentage":int}, engine='openpyxl')
        df.iloc[:,5] = df.iloc[:,5].str.rstrip()
        df.iloc[:,21] = df.iloc[:,21].str.rstrip()
        df.iloc[:,25] = df.iloc[:,25].str.rstrip()

        # Filter out "M FINANCIAL HOLDINGS INC" and "M HOLDINGS SECURITIES INC" from full["Agency Name"] and full["M Company Receiving Payment"]
        df = df[~df.iloc[:,5].isin(['M FINANCIAL HOLDINGS INC','M HOLDINGS SECURITIES INC'])]
        df = df[~df.iloc[:,21].isin(['M FINANCIAL HOLDINGS INC','M HOLDINGS SECURITIES INC'])]
        df = df[(df.iloc[:,25] == 'First Year Commission')] # Only want First Year Premium

        # Excluding policies that have duration greater than 1
        df["Duration"] = df["Duration"].fillna(0) 
        df = df[(df["Duration"].isin([0,1]))] 

        df.to_excel(os.path.join(self.datadir,str(month).zfill(2),f'Symetra{year*100+month}YTDfiltered.xlsx'),index=0)
        LOGGER.info(f'1680 Symetra: Symetra{year*100+month}YTD.xlsx updated.')

        df["SourceFileName"] = f"P-1680-LIF-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = ""
        df["CarrierContracteeID"] = df['Agency Number'].str.lstrip()
        df["CarrierContracteeName"] = df['Agency Name'].str.rstrip()
        df["ProductID"] = 0
        df["CarrierProductID"] = df['Product Code'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = ""
        df["CarrierProducerID"] = df['Agent Number'].str.rstrip()
        df["CarrierProducerName"] = df['Agent Name'].str.rstrip()
        df["PolicyNumber"] = df['Policy Number'].str.lstrip()
        df["IssueDate"] = df['Issue Date'].str[:4]+"/"+df['Issue Date'].str[4:6]+"/"+df['Issue Date'].str[6:8]
        df["InsuredName"] = df['Insured Name'].str.rstrip()
        df["YTDAnnualizedPrem"] = df['Paid Target Premium'].fillna(0) #.astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df['Paid Excess Premium'].fillna(0) #.astype(float).round(2).map("{:,.2f}".format)
        df["YTDFace"] = ""
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = ""
        df["SplitPercentage"] = 0.01*df["Producer Policy Split Percentage"]
        df["Replacement"] = ""
        df["CarrierProductName"] = df['Product Name'].str.rstrip()
        df["PolicyOwner"] = ""
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""    
        df = df[self.exportCols]
        

        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD = {excess}')
        
        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('Sym_lif_target').value = target
            notesbreakdown.range('Sym_lif_excess').value = excess
            notesbreakdown.range('Sym_lif_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('Sym_lif_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)
        LOGGER.info(f'{self.csvname} is saved at {self.csvdst}.')
        

def processSym(year, month):
    '''To generate a YTD report, if report month's data is not available, previous month's data would be used 
    temporarily until that month's data is arrived.'''
    sym = symetra(year, month)    
    isFileFound = sym.getrawmonth(year, month)
    if isFileFound == True:
        sym.aggregateYTD(year, month)
        sym.getcsv(year, month)
    else:
        try:
            if month == 1:
                sym.getcsv(year-1, 12)
            else:
                sym.getcsv(year,month -1)
        except FileNotFoundError:
            LOGGER.warning(f'Symetra {month_name[month-1]} file not available.')

# if __name__ == '__main__':
    # processSym(2022, 11)















########################################################################
    # def _recordCount(row,col,sheet): 
    #     # row col are the cell index of the cell start counting
    #     rowcount = 0
    #     for r in range(row,sheet.max_row+1,1):
    #         if sheet.cell(r,col).value != None:
    #             rowcount += 1 
    #         else:
    #             break
    #     return(rowcount)
