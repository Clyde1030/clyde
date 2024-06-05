import pandas as pd
from calendar import month_name, monthrange
import openpyxl as xl # report sales
import xlwings as xw
from xlwings.utils import rgb_to_int
import logging
import shutil
import os
import re

LOGGER = logging.getLogger(__name__)
# LOGGER = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
# handler = logging.StreamHandler()
# handler.setFormatter(formatter)
# LOGGER.addHandler(handler)
# LOGGER.setLevel(logging.DEBUG)

class nationwide:
    def __init__(self, year, month, *csvdst):
        self.year = year
        self.month = month
        self.monthApplied = str(month).zfill(2)
        # self.data = data
        self.csvdst = csvdst or f'C:\dev\Production\data\{year}\{self.monthApplied}'
        self.carrierId = 1490
        self.regex = re.compile(f'{self.monthApplied}.{year} M Financial Prod Report.xlsx') 
        self.datadir = r'J:\Acctng\Production\{}\Data\NW'.format(year)
        self.goanywhere =  r'J:\Acctng\Revenue\goanywhere\Nationwide'
        self.flash = os.path.join(f'C:\dev\Production',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        # self.flash = os.path.join(f'J:\Acctng\Production\{year}\Summary - Flash',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        self.csvnamelif = f"P-1490-LIF-{year}-{self.monthApplied}-LIF.txt"
        self.csvnamecoli = f"P-1490-LIf-{year}-{self.monthApplied}-COLI.txt"
        self.csvnameann = f"P-1490-ANN-{year}-{self.monthApplied}.txt"
        self.csvname401 = f'P-1490-401-{year}-{self.monthApplied}.txt'
        self.exportCols = ["SourceFileName","CarrierID","MemFirmID","CarrierContracteeID","CarrierContracteeName",
                            "ProductID","CarrierProductID","YearApplied","MonthApplied","ProducerID",
                            "CarrierProducerID","CarrierProducerName","PolicyNumber","IssueDate","InsuredName",
                            "YTDAnnualizedPrem" ,"YTDAnnualizedLowNon","YTDFace","RiskNumber","RiskName",
                            "Exchange1035","SplitPercentage","Replacement","CarrierProductName","PolicyOwner",
                            "Exchange","NLG","Modco","YRT","CSO2001"]

    def getrawmonth(self, year=None, month=None):
        '''Nationwide sends YTD production report through GoAnywhere every month. This fuction fetches the corresponding month's
        raw data from Goanywhere. Parameters year and month is default to self.year and self.month if not specified'''

        isFileFound = True

        if not year:
            year = self.year
        if not month:
            month = self.month
        regex = re.compile(f'{str(month).zfill(2)}.{year} M Financial Prod Report.xlsx')
        
        LOGGER.info(f'Searching {month_name[month]} file for carrierid {self.carrierId}...')

        f = []
        # If processing month is December, raw data will be delivered in next year January
        if month == 12:
            for dirs, _, filenames in os.walk(os.path.join(self.goanywhere,f'{year+1}','01')):
                f.extend([os.path.join(dirs,f) for f in filenames if re.search(regex,os.path.join(dirs,f))])            
        # Otherwise, raw data will be delivered in the next month.
        else:
            for dirs, _, filenames in os.walk(os.path.join(self.goanywhere,f'{year}',str(month+1).zfill(2))):
                f.extend([os.path.join(dirs,f) for f in filenames if re.search(regex,os.path.join(dirs,f))])        
        
        if len(f) == 0:
            LOGGER.info(f'File not available for {month_name[self.month]} yet. Go back to check Goanywhere folder.')
            path = None
            isFileFound = False
        elif len(f) == 1:
            path = f[0]            
        else: 
            path = f[len(f)-1] # To find the latest file

        if path != None:
            if os.path.exists(os.path.join(self.datadir,os.path.basename(path)))==False:
                shutil.copy2(path,self.datadir)
                LOGGER.info(f'{os.path.basename(path)} is copied to {month_name[self.month]} data directory')
            else:
                LOGGER.info(f'{os.path.basename(path)} has already existed in data directory.')            
        
        return isFileFound


    def _copyRange(self,startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected

    #Paste range
    #Paste data from copyRange into template sheet
    def _pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def _recordCount(self,row,col,sheet): 
        # row col are the cell index of the cell start counting
        rowcount = 0
        for r in range(row,sheet.max_row+1,1):
            if sheet.cell(r,col).value != None:
                rowcount += 1 
            else:
                break
        return rowcount

    def getcsvlif(self, datayear=None, datamonth=None):
        '''This function generates csv files for individual life business. datayear and datamonth are 
        provided if a different month's data is reported as default month'''

        if not datayear:
            datayear = self.year
        if not datamonth:
            datamonth = self.month

        regex = re.compile(f'{str(datamonth).zfill(2)}.{datayear} M Financial Prod Report.xlsx')
        file = list(filter(regex.match,os.listdir(self.datadir)))[0]

        LOGGER.info(f'Begin Processing Nationwide Individual life with {datayear} {month_name[datamonth]} data...')
        wb = xl.Workbook()
        life = wb.active        
        recordCounts = [0]
        l = ['TL','UL','VL']
        for i in range(len(l)):
            sht = xl.load_workbook(os.path.join(self.datadir,file))[l[i]]
            cnt = self._recordCount(6,1,sht)
            datacopying = self._copyRange(1,6,24,5+cnt,sht)
            self._pasteRange(1,sum(recordCounts)+1,24,sum(recordCounts)+cnt,life,datacopying)
            recordCounts.append(cnt)
        life.insert_rows(1)
        headers = ["AgentNumber","AgentName","AgencyNumber","ProductCarrier","AgencyName",
            "PolicyNumber","Mode","ProductType","ProductGroup","ProductName",
            "ProductCode","InsuredName","IssueDate","MonthEndDate","AnnualizedPremium",
            "ExcessPremium","FaceAmount","ProducerShare","PolicyCount","1035",
            "InternalExchange","2001CSOCompliant","ModCo","NoLapseGuarantee"]
        for i in range(0,len(headers),1):
            life.cell(1,i+1).value = headers[i]
        wb.save(os.path.join(self.datadir,'lifetemp.xlsx'))            

        df = pd.read_excel(os.path.join(self.datadir,'lifetemp.xlsx'),converters={'AgencyNumber':str,'ProductCode':str}, engine='openpyxl')

        df["SourceFileName"] = f"P-1490-LIF-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = ""
        df["CarrierContracteeID"] = df['AgencyNumber'].astype(str)
        df["CarrierContracteeName"] = df['AgencyName'].str.rstrip().str[0:50]
        df["ProductID"] = 0
        df["CarrierProductID"] = df.apply(lambda x: x["ProductName"] if x['ProductCode']=='.' else x["ProductCode"], axis=1) #df['ProductCode'].astype(str)
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['AgentNumber'].astype(str)
        df["CarrierProducerName"] = df['AgentName'].str.rstrip().str[0:50]
        df["PolicyNumber"] = df['PolicyNumber'].astype(str)
        # df["IssueDate"] = df['IssueDate']
        df["InsuredName"] = df['InsuredName']
        df["YTDAnnualizedPrem"] = df['AnnualizedPremium']
        df["YTDAnnualizedLowNon"] = df['ExcessPremium']
        df["YTDFace"] = df['FaceAmount'].astype(float).round(2).map("{:,.2f}".format)
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = df['1035'].str[0:1]
        df["SplitPercentage"] = df['PolicyCount']
        df["Replacement"] = ""
        df["CarrierProductName"] = df['ProductName']
        df["PolicyOwner"] = ""
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = df["2001CSOCompliant"].str[0:1]    


        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        TLTarget = df.loc[df["CarrierProductName"].str.contains('TERM',case = False)==True,"YTDAnnualizedPrem"].sum()
        TLExcess = df.loc[df["CarrierProductName"].str.contains('TERM',case = False)==True,"YTDAnnualizedLowNon"].sum()
        VLTarget = df.loc[df["ProductType"].str.contains('VARIABLE',case = False)==True,"YTDAnnualizedPrem"].sum()
        VLExcess = df.loc[df["ProductType"].str.contains('VARIABLE',case = False)==True,"YTDAnnualizedLowNon"].sum()
        ULTarget = target - TLTarget - VLTarget
        ULExcess = excess - TLExcess - VLExcess
        CaremattersTarget = df.loc[df["CarrierProductName"].str.contains('CAREMATTERS',case = False)==True,"YTDAnnualizedPrem"].sum()
        CaremattersExcess = df.loc[df["CarrierProductName"].str.contains('CAREMATTERS',case = False)==True,"YTDAnnualizedLowNon"].sum()

        LOGGER.info(f'Target Premium {month_name[self.month]} YTD TL = {TLTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD TL = {TLExcess}')
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD VL = {VLTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD VL = {VLExcess}')
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD UL = {ULTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD UL = {ULExcess}')
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD LTC = {CaremattersTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD LTC = {CaremattersExcess}')

        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('NW_lif_TL_target').value = TLTarget
            notesbreakdown.range('NW_lif_TL_excess').value = TLExcess
            notesbreakdown.range('NW_lif_UL_target').value = ULTarget
            notesbreakdown.range('NW_lif_UL_excess').value = ULExcess
            notesbreakdown.range('NW_lif_VL_target').value = VLTarget
            notesbreakdown.range('NW_lif_VL_excess').value = VLExcess
            notesbreakdown.range('NW_LTC_target').value = CaremattersTarget
            notesbreakdown.range('NW_LTC_excess').value = CaremattersExcess
            notesbreakdown.range('NW_lif_TL_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_lif_TL_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_lif_UL_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_lif_UL_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_lif_VL_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_lif_VL_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_LTC_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_LTC_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvnamelif), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)        
        LOGGER.info(f'{self.csvnamelif} is saved at {self.csvdst}.')
        os.remove(os.path.join(self.datadir,'lifetemp.xlsx'))



    def getcsvcoli(self, datayear=None, datamonth=None):
        '''This function generates csv files for COLI life business. datayear and datamonth are 
        provided if a different month's data is reported as default month'''

        if not datayear:
            datayear = self.year
        if not datamonth:
            datamonth = self.month

        regex = re.compile(f'{str(datamonth).zfill(2)}.{datayear} M Financial Prod Report.xlsx')
        file = list(filter(regex.match,os.listdir(self.datadir)))[0]

        LOGGER.info(f'Begin Processing Nationwide COLI life with {datayear} {month_name[datamonth]} data...')

        # Coli data lives in the same worbook but different worksheets in the excel file Nationwide sends us every month.
        # Locate the file name of the raw data file in the data directory  
        # We want both COLI-REG and COLI-PP worksheets
        reg = pd.read_excel(os.path.join(self.datadir, file), sheet_name='COLI-REG', converters = {'AGENT_NUMBER':str,'AGENCY_NUMBER':str, 'PRODUCT_CODE':str}, engine='openpyxl')
        pp = pd.read_excel(os.path.join(self.datadir, file), sheet_name='COLI-PP', converters = {'AGENT_NUMBER':str,'AGENCY_NUMBER':str, 'PRODUCT_CODE':str}, engine='openpyxl')

        df = pd.concat([reg,pp], ignore_index=True,sort = False)
        df = df.loc[pd.isnull(df['POLICY_NUMBER'])==False]
        df["SourceFileName"] = f"P-1490-LIF-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = 1490
        df["MemFirmID"] = ""
        df["CarrierContracteeID"] = df['AGENCY_NUMBER'].str.rstrip()
        df["CarrierContracteeName"] = df['AGENCY_NAME'].str.rstrip()
        df["ProductID"] = 0
        df["CarrierProductID"] = df['PRODUCT_CODE'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['AGENT_NUMBER'].str.rstrip()
        df["CarrierProducerName"] = df['AGENT_NAME'].str.rstrip()
        df["PolicyNumber"] = df['POLICY_NUMBER']
        df["IssueDate"] = df['ISSUE_DATE']
        df["InsuredName"] = df['INSURED'].str.rstrip()
        df["IsPrivatePlacement"] = df['PRODUCT_NAME'].str.contains('PP')
        df["YTDAnnualizedPrem"] = df.apply(lambda x: 0 if x['PRODUCT_NAME']=='NWL PPVUL' else x['ANNUALIZED_PREMIUM'], axis=1) 
        df["YTDAnnualizedLowNon"] = df.apply(lambda x: x['ANNUALIZED_PREMIUM'] if x['PRODUCT_NAME']=='NWL PPVUL' else x['EXCESS_PREMIUM'], axis=1) 
        df["YTDFace"] = df['FACE_AMOUNT'].astype(float).round(2).map("{:,.2f}".format)
        df["RiskNumber"] = df["CASE_NUMBER"].str.rstrip()
        df["RiskName"] = df["CASE_NAME"].str.rstrip()
        df["Exchange1035"] = df['TEN_35'].str[0:1]
        df["SplitPercentage"] = df['POLICY_COUNT']
        df["Replacement"] = ""
        df["CarrierProductName"] = df['PRODUCT_NAME'].str.rstrip()
        df["PolicyOwner"] = ""
        df["Exchange"] = df['INTERNAL/EXTERNAL']
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ''    


        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        ColiRegTarget = df.loc[df["CarrierProductName"] == 'NWL',"YTDAnnualizedPrem"].sum()
        ColiRegExcess = df.loc[df["CarrierProductName"] == 'NWL',"YTDAnnualizedLowNon"].sum()
        ColiPPTarget = df.loc[df["CarrierProductName"].str.contains('NWL PPVUL',case = False)==True,"YTDAnnualizedPrem"].sum()
        ColiPPExcess = df.loc[df["CarrierProductName"].str.contains('NWL PPVUL',case = False)==True,"YTDAnnualizedLowNon"].sum()

        LOGGER.info(f'Target Premium {month_name[self.month]} YTD Coli Reg = {ColiRegTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD Coli Reg = {ColiRegExcess}')
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD Coli PP = {ColiPPTarget}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD Coli PP = {ColiPPExcess}')

        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('NW_Coli_Reg_target').value = ColiRegTarget
            notesbreakdown.range('NW_Coli_Reg_excess').value = ColiRegExcess
            notesbreakdown.range('NW_Coli_PP_target').value = ColiPPTarget
            notesbreakdown.range('NW_Coli_PP_excess').value = ColiPPExcess
            notesbreakdown.range('NW_Coli_Reg_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_Coli_Reg_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_Coli_PP_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_Coli_PP_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvnamecoli), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)        
        LOGGER.info(f'{self.csvnamecoli} is saved at {self.csvdst}.')

    def getcsv401(self, datayear=None, datamonth=None):
        '''This function generates csv files for COLI life business. datayear and datamonth are 
        provided if a different month's data is reported as default month'''

        if not datayear:
            datayear = self.year
        if not datamonth:
            datamonth = self.month

        regex = re.compile(f'{str(datamonth).zfill(2)}.{datayear} M Financial Prod Report.xlsx')
        file = list(filter(regex.match,os.listdir(self.datadir)))[0]

        LOGGER.info(f'Begin Processing Nationwide 401K with {datayear} {month_name[datamonth]} data...')
        df = pd.read_excel(os.path.join(self.datadir,file),sheet_name='Pensions',header=0, engine='openpyxl')
        df = df.head(df.shape[0] -1)

        df["SourceFileName"] = f"P-1490-401-{self.year}-{self.monthApplied}-1"
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = df['AGENCY NUMBER']
        df["CarrierContracteeName"] = df['AGENCY NAME'].str.rstrip()
        df["ProductID"] = 0
        df["CarrierProductID"] = df['PRODUCT CODE'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['AGENT APPT NUMBER']
        df["CarrierProducerName"] = df['AGENT NAME']
        df["PolicyNumber"] = df['POLICY NUMBER']
        df["IssueDate"] = df['ISSUE DATE']
        df["InsuredName"] = df['INSURED NAME'].str.rstrip()
        df["YTDAnnualizedPrem"] = df['ANNUALIZED TARGET PREMIUM'] #.astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = 0
        df["YTDFace"] = df['CURRENT MONTH ASSET VALUE'].astype(float).round(2).map("{:,.2f}".format)
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = 'N'
        df["SplitPercentage"] =  df.apply(lambda x: x['POLICY COUNT']*0.01 if pd.isnull(x['POLICY COUNT'])==False else x['POLICY COUNT'], axis = 1)
        df["Replacement"] = ""
        df["CarrierProductName"] = df['PRODUCT NAME']
        df["PolicyOwner"] = ""
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""
        df["year"] = pd.DatetimeIndex(df['IssueDate']).year

        df = df[df['year'] >= self.year - 1]

        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD Pension = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD Pension = {excess}')

        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('NW_401_target').value = target
            notesbreakdown.range('NW_401_excess').value = excess
            notesbreakdown.range('NW_401_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_401_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname401), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)
        LOGGER.info(f'{self.csvname401} is saved at {self.csvdst}.')


    def getcsvann(self, datayear=None, datamonth=None):
        '''This function generates csv files for COLI life business. datayear and datamonth are 
        provided if a different month's data is reported as default month'''

        if not datayear:
            datayear = self.year
        if not datamonth:
            datamonth = self.month

        regex = re.compile(f'{str(datamonth).zfill(2)}.{datayear} M Financial Prod Report.xlsx')
        file = list(filter(regex.match,os.listdir(self.datadir)))[0]

        LOGGER.info(f'Begin Processing Nationwide Annuity life with {datayear} {month_name[datamonth]} data...')
        df = pd.read_excel(os.path.join(self.datadir,file),sheet_name='IA Sales',header=2, engine='openpyxl')
        df = df.head(df.shape[0] -1)
        df["SourceFileName"] = f"P-1490-ANN-{self.year}-{self.monthApplied}-1"
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = df['Agency No']
        df["CarrierContracteeName"] = df['Agency Name'].str.rstrip()
        df["ProductID"] = 0
        df["CarrierProductID"] = df['Product Code'].str.rstrip().fillna(0)
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df['Agent Number']
        df["CarrierProducerName"] = df['Agent Name']
        df["PolicyNumber"] = df['Policy Number']
        df["IssueDate"] = df['Issue Date']
        df["InsuredName"] = df['Insured Name']
        df["YTDAnnualizedPrem"] = df['Target Premium'] #.astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df['Excess Premium'] #.astype(float).round(2).map("{:,.2f}".format)
        df["YTDFace"] = df['Face Amount'].astype(float).round(2).map("{:,.2f}".format)
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = df['1035']
        df["SplitPercentage"] = df.apply(lambda x: 0 if pd.isnull(x['Producer Share'])==True else x['Producer Share'], axis = 1).astype(int)*0.01   
        df["Replacement"] = ""
        df["CarrierProductName"] = df['Product Name']
        df["PolicyOwner"] = ""
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""
  

        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD Annuity = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD Annuity = {excess}')

        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('NW_Ann_target').value = target
            notesbreakdown.range('NW_Ann_excess').value = excess
            notesbreakdown.range('NW_Ann_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('NW_Ann_excess').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvnameann), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)        
        LOGGER.info(f'{self.csvnameann} is saved at {self.csvdst}.')




def processNW(year, month):
    '''To generate a YTD report, if report month's data is not available, previous month's data would be used 
    temporarily until that month's data is arrived.'''
    nw = nationwide(year,month)
    isFileFound = nw.getrawmonth(year, month)
    if isFileFound == True:
        nw.getcsvlif(year, month)
        nw.getcsvcoli(year, month)
        nw.getcsv401(year, month)
        nw.getcsvann(year, month)
    else:
        try:
            if month == 1:
                nw.getcsvlif(year-1, 12)
                nw.getcsvcoli(year-1, 12)
                nw.getcsv401(year-1, 12)
                nw.getcsvann(year-1, 12)
            else:
                nw.getcsvlif(year,month -1)
                nw.getcsvcoli(year,month -1)
                nw.getcsv401(year,month -1)
                nw.getcsvann(year,month -1)
        except IndexError:
            LOGGER.warning(f'Nationwide {month_name[month-1]} file not available.')


# if __name__ == '__main__':
#     processNW(2022,11)

        

















###############################

## Code Archive

        # This function generates csv files for COLI business.

        # Coli data lives in the same worbook but different worksheets in the excel file Nationwide sends us every month.
        # Locate the file name of the raw data file in the data directory  
        # file = list(filter(self.regex.match,os.listdir(self.datadir)))[0]
        # LOGGER.debug(f'parsing Nationwide COLI raw file...')

        # Create a workbook to aggregate COLI data from the raw data and save as 'colitemp.xlsx' in the data directory
        # wb = xl.Workbook()
        # life = wb.active        
        # recordCounts = [0]
        # We want COLI-REG and COLI-PP worksheets
        # l = ['COLI-REG','COLI-PP']

        # Move data to 'colitemp.xlsx'
        # for i in range(len(l)):
        #     sht = xl.load_workbook(os.path.join(self.datadir,file))[l[i]]
        #     cnt = self._recordCount(2,1,sht)
        #     datacopying = self._copyRange(1,2,24,1+cnt,sht)
        #     self._pasteRange(1,sum(recordCounts)+1,24,sum(recordCounts)+cnt,life,datacopying)
        #     recordCounts.append(cnt)
        # life.insert_rows(1)
        # headers = ['AGENT_NUMBER','AGENT_NAME','AGENCY_NUMBER','AGENCY_NAME','POLICY_NUMBER',
        #             'PRODUCT_TYPE','PRODUCT_NAME','PRODUCT_CODE','INSURED','ISSUE_DATE',
        #             'MONTH_END_DATE','ANNUALIZED_PREMIUM','EXCESS_PREMIUM','FACE_AMOUNT','PRODUCER_COMMISSION_RATE',
        #             'POLICY_COUNT','1035','INTERNAL/EXTERNAL','GROUP_NUMBER','CASE_NUMBER','CASE_NAME',
        #             'ADD_TO_FILE_DATE','YTD_PREMIUM','ADDL_PROTECTION_RIDER']
        # for i in range(0,len(headers),1):
        #     life.cell(1,i+1).value = headers[i]
        # wb.save(os.path.join(self.datadir,'colitemp.xlsx'))            

        # Use pandas to read in 'colitemp.xlsx' data and keep on processing
        # df = pd.read_excel(os.path.join(self.datadir,'colitemp.xlsx'),converters = {'AGENT_NUMBER':str,'AGENCY_NUMBER':str, 'PRODUCT_CODE':str})



        # os.remove(os.path.join(self.datadir,'colitemp.xlsx'))


