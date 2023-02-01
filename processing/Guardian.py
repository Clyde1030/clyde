import pandas as pd
from calendar import month_name, monthrange
import openpyxl as xl 
import xlwings as xw
from xlwings.utils import rgb_to_int
import logging
import os
import re

LOGGER = logging.getLogger(__name__)
# LOGGER = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
# handler = logging.StreamHandler()
# handler.setFormatter(formatter)
# LOGGER.addHandler(handler)
# LOGGER.setLevel(logging.DEBUG)

class guardian:

    def __init__(self, year, month, *csvdst):
        self.year = year
        self.month = month
        self.monthApplied = str(month).zfill(2)
        # self.data = data
        self.carrierId = 1330 
        self.regex = re.compile('M Financial_\d{6}.xlsx')  
        self.datadir = f'J:\Acctng\Production\{self.year}\Data\Guardian'
        self.goanywhere = None
        self.csvname = f"P-{self.carrierId}-DIS-{year}-{self.monthApplied}.txt"
        self.receiveremail = 'yu-sheng.lee@mfin.com'
        self.csvdst = csvdst or f'C:\dev\Production\data\{self.year}\{self.monthApplied}'
        self.flash = os.path.join(f'C:\dev\Production',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        # self.flash = os.path.join(f'J:\Acctng\Production\{year}\Summary - Flash',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        self.exportCols = ["SourceFileName","CarrierID","MemFirmID","CarrierContracteeID","CarrierContracteeName",
                            "ProductID","CarrierProductID","YearApplied","MonthApplied","ProducerID",
                            "CarrierProducerID","CarrierProducerName","PolicyNumber","IssueDate","InsuredName",
                            "YTDAnnualizedPrem" ,"YTDAnnualizedLowNon","YTDFace","RiskNumber","RiskName",
                            "Exchange1035","SplitPercentage","Replacement","CarrierProductName","PolicyOwner",
                            "Exchange","NLG","Modco","YRT","CSO2001"]
    
    def _copyRange(self,startCol, endCol, startRow, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
        return rangeSelected

    def _pasteRange(self, startCol, endCol, startRow, endRow, sheetReceiving,copiedData):
    #Paste data from copyRange into template sheet
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def _ytdlist(self, month=None):
        if not month:
            month = self.month
        ytdlist = []
        path = []
        for i in range(1,month+1):
            if i <= month:
                ytdlist.append(f'M Financial_{self.year}{str(i).zfill(2)}.xlsx') 
        path= [os.path.join(self.datadir,i) for i in os.listdir(self.datadir) if i in ytdlist]

        LOGGER.debug(f'{len(path)} raw data files are appended to the list.')
        return path

    def isFileFound(self, year=None, month=None):

        if not year:
            year = self.year
        if not month:
            month = self.month

        isFileFound = True
        monthfilename = f'M Financial_{year}{str(month).zfill(2)}.xlsx'

        if os.path.exists(os.path.join(self.datadir,monthfilename)):
            LOGGER.info(f'{monthfilename} exists')
        else:
            LOGGER.info(f'{monthfilename} does not exists')
            isFileFound = False
        return isFileFound

    def aggregateYTD(self):
        
        # Create a file list that are override files and needed to be aggregated
        filelist = self._ytdlist()
        LOGGER.debug(f'Start combining {len(filelist)} files in {self.datadir}...')
        headers = ['SourceFileName','AgentWritingCode','ProducerIdentifier','FullName','PolicyNumber',
            	   'ACTTransactActivityDate','modeType','PolicyDuration','PolicyEffectiveDate','PaymentDueDate',
                   'InsuredFullName','CompensationCategory1','CaseNumber','CaseName','CommissionAmount',
                   'CommissionRate','CommissionRate2','CommissionAmount2','CommissionablePremium']
        ytd = xl.Workbook()
        full = ytd.active
        rowcounts = [0] # contains the row counts where each override worksheet will be copy from

        # Append each month's file to YTD workbook        
        for i in filelist:
            LOGGER.info(f'now processing {i}...')
            filename = os.path.basename(i)
            processingSheet = xl.load_workbook(i).worksheets[0]
            rowcount = len(processingSheet['A'])-1
            datapCopying = self._copyRange(1, 18, 2, 1+rowcount, sheet=processingSheet)
            self._pasteRange(2,19,1+sum(rowcounts),rowcount+sum(rowcounts), full, datapCopying)
            for row in range(sum(rowcounts)+1,sum(rowcounts)+rowcount+1,1):      
                full.cell(row,1).value = filename
            rowcounts.append(rowcount)
        
        # Fill the headers and apply filter
        full.insert_rows(1)
        for i, item in enumerate(headers):
            full.cell(1,1+i).value = item
        full.auto_filter.ref = f'A1:S{sum(rowcounts)+1}' 

        ytd.save(os.path.join(self.datadir,'GuardianYTD.xlsx'))
        LOGGER.info(f'1330 Guardian: GuardianYTD.xlsx generated.')        


    def getcsv(self):

        LOGGER.info(f'Processing 1330 Guardian Disability with {month_name[self.month]} YTD data...')

        df = pd.read_excel(os.path.join(self.datadir,'GuardianYTD.xlsx'),converters={"ProducerIdentifier":str}, engine='openpyxl')
        
        # Take only first year premium as sales. NO RENEWALS!!
        df = df[df['CompensationCategory1'].str.rstrip() == 'FIRST YEAR']      
        
        df["SourceFileName"] = f"P-{self.carrierId}-DIS-{self.year}-{self.monthApplied}-1"
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = df["FullName"].str.rstrip()
        df["CarrierContracteeName"] = df["FullName"].str.rstrip()
        df["ProductID"] = 0
        df["CarrierProductID"] = "Provider Choice"
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df["ProducerIdentifier"].str.rstrip()
        df["CarrierProducerName"] = df["FullName"].str.rstrip()
        df["PolicyNumber"] = df["PolicyNumber"].str.rstrip()
        df["IssueDate"] = df["PolicyEffectiveDate"].dt.normalize()
        df["InsuredName"] = df["InsuredFullName"].str.rstrip()
        df["YTDAnnualizedPrem"] = df["CommissionablePremium"]
        df["YTDAnnualizedLowNon"] = 0
        df["YTDFace"] = ""
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = ""
        df["SplitPercentage"] = df["CommissionRate2"]
        df["Replacement"] = ""
        df["CarrierProductName"] = "Provider Choice"
        df["PolicyOwner"] = df["CaseName"].str.rstrip()
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""


        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD = {excess}')
        
        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']         
            notesbreakdown.range('Guar_target').value = target # notesbreakdown['C110'].value = target
            notesbreakdown.range('Guar_target').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df["YTDAnnualizedPrem"].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df["YTDAnnualizedLowNon"].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname), sep='|', index=False)        
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)        
        LOGGER.info(f'{self.csvname} is saved at {self.csvdst}.')

def processGuard(year, month):
    guard = guardian(year,month)
    if guard.isFileFound()==False:
        LOGGER.info(f'Guardian {month_name[guard.month]} file is not available yet. Process Guardian with available YTD data...')
    guard.aggregateYTD()
    guard.getcsv()

# if __name__ == '__main__':
#     processGuard(2022,11)


