import openpyxl as xl
from calendar import month_name
import xlwings as xw
from xlwings.utils import rgb_to_int
import pandas as pd
import logging
import re
import os

LOGGER = logging.getLogger(__name__)
# LOGGER = logging.getLogger()
# formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(filename)s - %(message)s', '%Y-%m-%d %H:%M:%S')
# handler = logging.StreamHandler()
# handler.setFormatter(formatter)
# LOGGER.addHandler(handler)
# LOGGER.setLevel(logging.DEBUG)

class metlife:

    def __init__(self, year, month, *csvdst):
        self.year = year
        self.month = month
        self.monthApplied = str(month).zfill(2)
        self.carrierId = 1325
        self.regex = re.compile("M[_-]GROUP.*(.xlsx)$",re.IGNORECASE) 
        self.datadir = f'J:\Acctng\Production\{self.year}\Data\MetLife'
        self.csvname = f'P-1325-DIS-{year}-{self.monthApplied}-1.txt'
        self.csvdst = csvdst or f'C:\dev\Production\data\{year}\{self.monthApplied}'
        self.flash = os.path.join(f'C:\dev\Production',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        # self.flash = os.path.join(f'J:\Acctng\Production\{year}\Summary - Flash',f'Production Summary {year}-{self.monthApplied}.xlsm')        
        self.goanywhere = None
        self.exportCols = ["SourceFileName","CarrierID","MemFirmID","CarrierContracteeID","CarrierContracteeName",
                            "ProductID","CarrierProductID","YearApplied","MonthApplied","ProducerID",
                            "CarrierProducerID","CarrierProducerName","PolicyNumber","IssueDate","InsuredName",
                            "YTDAnnualizedPrem" ,"YTDAnnualizedLowNon","YTDFace","RiskNumber","RiskName",
                            "Exchange1035","SplitPercentage","Replacement","CarrierProductName","PolicyOwner",
                            "Exchange","NLG","Modco","YRT","CSO2001"]


    def _copyRange(self, startCol, endCol, startRow, endRow, sheet):
        rangeSelected = []
        #Loops through selected Rows
        for i in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            for j in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = i, column = j).value)
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
    
        return rangeSelected


    #Paste range
    #Paste data from copyRange into template sheet
    def _pasteRange(self, startCol, endCol, startRow, endRow, sheetReceiving,copiedData):
        countRow = 0
        for i in range(startRow,endRow+1,1):
            countCol = 0
            for j in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1

    def isFileFound(self, year=None, month=None):
        if not year:
            year = self.year
        if not month:
            month = self.month
        isFileFound = True
        monthfilename = f'M-GROUP {year}{str(month).zfill(2)}.xlsx'
        if os.path.exists(os.path.join(self.datadir,monthfilename)):
            LOGGER.info(f'{monthfilename} exists')
        else:
            LOGGER.info(f'{monthfilename} does not exists')
            isFileFound = False
        return isFileFound


    def _ytdlist(self, month=None):
        if not month:
            month = self.month
        # ytdmonth = []
        path = []
        # for i in range(1,month+1):
        #     if i <= month:
        #         ytdmonth.append(str(i).zfill(2))
        for dirs, _, filenames in os.walk(self.datadir): 
            files = [os.path.join(dirs,f) for f in filenames if re.search(self.regex,os.path.join(dirs,f))]
            path.extend(files)
        LOGGER.debug(f'1325 MetLife: {len(path)} raw data files are appended to the list.')
        return path

    def aggregateYTD(self):
        filelist = self._ytdlist()
        LOGGER.debug(f'Start combining {len(filelist)} files in the list...')
        headers = ["Source File Name","EFT Statement Date","Payee Broker Code","Policy #","Group #",
                    "Company Name","Bill Date","Insured Name","Effective Date","Form Code",	
                    "Form Description","Form Edition Year","Elimination Period","Multi-Life Discount","Commission Type Description",
                    "Reason Description","Broker Name","Premium","Commission Rate%","Commission",
                    "SPLT_PCT","First Year/Renewal","Writing Producer 1 Name","Writing Producer 1 Split %","Writing Producer 2 Name",
                    "Writing Producer 2 Split %","Writing Producer 3 Name","Writing Producer 3 Split %","Writing Producer 4 Name","Writing Producer 4 Split %",
                    "Writing Producer 5 Name","Writing Producer 5 Split %","Sub-GA 1 Name","Sub-GA 1 Commission Rate %","Sub-GA 1 Commission",
                    "Sub-GA 2 Name","Sub-GA 2 Commission Rate %","Sub-GA 2 Commission","TRAN_RLS_DT"]
        Full = xl.Workbook()
        ws = Full.active
        rowcounts = [0]
        for i in filelist:
            LOGGER.info(f'now processing {i}...')
            filename = os.path.basename(i)
            sheet = xl.load_workbook(i).worksheets[0]
            rowcount = len(sheet['A'])-2
            datacopying = self._copyRange(1,38,2,1+rowcount,sheet)
            self._pasteRange(2,39,1+sum(rowcounts),rowcount+sum(rowcounts),ws,datacopying)
            for row in range(sum(rowcounts)+1,sum(rowcounts)+rowcount+1,1):      
                ws.cell(row,1).value = filename
            rowcounts.append(rowcount)
        ws.insert_rows(1)
        for i, item in enumerate(headers):
            ws.cell(1,1+i).value = item 

        ws.auto_filter.ref = f'A1:AM{sum(rowcounts)+1}' 
        Full.save(os.path.join(self.datadir,f'MetLifeYTD.xlsx'))
        LOGGER.info(f'1325 MetLife: MetLifeYTD.xlsx generated.')

        
    def getcsv(self):
        
        LOGGER.debug(r'Parsing MetLife raw file...')
        df = pd.read_excel(os.path.join(self.datadir,'MetLifeYTD.xlsx'),converters={'Policy #':str,'Form Code':str,'Writing Producer 1 Split %':int}, engine='openpyxl')

        df = df[(df.iloc[:,21] == 'F')] # Only want First Year Premium
        df = df[(df["Commission Type Description"].str.contains('FIRST YEAR')==True)]
        df = df[(df["Broker Name"].str.rstrip()!='FRED HILL')]

        df["SourceFileName"] = f"P-1325-DIS-{self.year}-{self.monthApplied}-1" 
        df["CarrierID"] = self.carrierId
        df["MemFirmID"] = 0
        df["CarrierContracteeID"] = ""
        df["CarrierContracteeName"] = ""
        df["ProductID"] = 0
        df["CarrierProductID"] = df['Form Code'].str.rstrip()
        df["YearApplied"] = self.year
        df["MonthApplied"] = self.monthApplied
        df["ProducerID"] = 0
        df["CarrierProducerID"] = df.apply(lambda x: x['Broker Name'] if (x['Writing Producer 1 Name'] == '' or x['Writing Producer 1 Name'] == ' ' or pd.isnull(x['Writing Producer 1 Name']) == True) else x['Writing Producer 1 Name'], axis=1)
        df["CarrierProducerID"] = df["CarrierProducerID"].str.rstrip()
        df["CarrierProducerName"] = df["CarrierProducerID"].str.rstrip()
        df["PolicyNumber"] = df['Policy #'].str.lstrip().str.rstrip()
        df["IssueDate"] = pd.to_datetime(df['Effective Date'], format='%m/%d/%Y').dt.normalize()
        df["InsuredName"] = df['Insured Name'].str.rstrip()
        df["YTDAnnualizedPrem"] = df['Premium'] 
        df["YTDAnnualizedLowNon"] = 0
        df["YTDFace"] = 0
        df["RiskNumber"] = ""
        df["RiskName"] = ""
        df["Exchange1035"] = ""
        df["SplitPercentage"] = df["Writing Producer 1 Split %"].fillna(0)*0.01
        df["Replacement"] = ""
        df["CarrierProductName"] = df['Form Description'].str.rstrip()
        df["PolicyOwner"] = ""
        df["Exchange"] = ""
        df["NLG"] = ""
        df["Modco"] = ""
        df["YRT"] = ""
        df["CSO2001"] = ""    

        # Report
        target = df["YTDAnnualizedPrem"].sum()
        excess = df["YTDAnnualizedLowNon"].sum()
        LOGGER.info(f'Target Premium {month_name[self.month]} YTD = {target}')
        LOGGER.info(f'Excess Premium {month_name[self.month]} YTD = {excess}')

        LOGGER.info(f'Accessing flash workbook and updating values...')
        with xw.App(visible=False) as app:
            flash = xw.Book(self.flash)
            notesbreakdown = flash.sheets['Notes-Breakdown']
            notesbreakdown.range('MetLife_target').value = target
            notesbreakdown.range('D76').value = excess
            notesbreakdown.range('MetLife_target').api.Font.Color = rgb_to_int((0, 102, 204))
            notesbreakdown.range('D76').api.Font.Color = rgb_to_int((0, 102, 204))
            flash.save(self.flash)
            flash.close()
        LOGGER.info(f'{self.flash} is updated and saved')

        # Export
        df["YTDAnnualizedPrem"] = df['YTDAnnualizedPrem'].astype(float).round(2).map("{:,.2f}".format)
        df["YTDAnnualizedLowNon"] = df['YTDAnnualizedLowNon'].astype(float).round(2).map("{:,.2f}".format)
        df = df[self.exportCols]
        df.to_csv(os.path.join(self.csvdst,self.csvname), index=False, sep = '|')
        # df.to_excel(os.path.join(self.csvdst,'test.xlsx'), index=False)
        LOGGER.info(f'{self.csvname} is saved at: {self.csvdst}')


def processmetL(year,month):
    met = metlife(year,month)
    met.aggregateYTD()
    met.getcsv()



# if __name__ == '__main__':
#     processmetL(2022,10)
